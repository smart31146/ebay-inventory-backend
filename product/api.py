import csv
import json
import datetime
import openpyxl
import environ
import os
from django.conf import settings
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from dry_rest_permissions.generics import DRYPermissions
from pathlib import Path
from ebaysdk.trading import Connection
from ebaysdk.shopping import Connection as Shopping
import psycopg2
from django.db.models import Q
from .models import Product, DeletedList, OrderList
from .serializers import ProductSerializer
from product.scrape.engineselector import select_engine
from utils.convertcurrency import getCurrentRate
from users.models import User
from utils.profit_formula import profit_formula
from utils.sell_set_formula import sell_set_formula
from utils.profit_order_formula import profit_order_formula

class ProductViewSet(ModelViewSet):
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    permission_classes = (DRYPermissions, )

    @action(detail=False, methods=['POST'])
    def scrape_data(self, request):
        url = request.data['url']
        engine = select_engine(url=url)
        
        if engine:
            engine = engine()
            try:
                data = engine.scrape_data(source_url=url)
                # data['price_en'] = convert('JPY', 'USD', data['price_jp'])

                return Response(
                    data=data,
                    status=200
                )
            except Exception as err:
                raise err
        
        return Response(
            data='入力したサイトへのサービスはまだサポートされていません。',
            status=400
        )
            
    @action(detail=False, methods=['POST'])
    def validate_product(self, request):
        info = request.data['product_info']
        url = info['url']
        mode = info['mode']

        if mode == 1:
            products = []

            try:
                # validate duplicate
                products = Product.objects.filter(Q(purchase_url=url) | Q(ebay_url=url), deleted=False).values()

            except Exception as err:

                return Response(
                    data = 'データベース接続失敗！',
                    status = 401
                )

            if len(products) > 0:
                return Response(
                    data = 'すでに存在しています！',
                    status = 401
                )

            return Response(
                data = True,
                status = 200
            )

        # engine = select_engine(url = purchase_url)
        
        # if engine:
        #     engine = engine()
        #     try:
        #         data = engine.scrape_data(source_url = purchase_url)

        #         if data['nothing'] == False:
        #             return Response(
        #                 data = data,
        #                 status = 200
        #             )
        #         else:
        #             return Response(
        #                 data = 'この商品は削除されました。',
        #                 status = 401
        #             )
        #     except:
        #         return Response(
        #             data = '入力したサイトへのサービスはまだサポートされていません。',
        #             status = 401
        #         )
        # else:
        #     return Response(
        #         data = '入力したサイトへのサービスはまだサポートされていません。',
        #         status = 401
        #     )
        
    @action(detail=False, methods=['GET'])
    def get_products(self, request):
        creator = request.GET.get('created_by')
        user_id = request.GET.get('user_id')
        superuser = self.isSuperUser(user_id)

        products = []

        if creator == '':
            if superuser == True:
                products = Product.objects.filter(deleted = False).select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'notes', 'created_by_id')
        else:
            products = Product.objects.filter(created_by = creator, deleted = False).select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'notes', 'created_by_id')
        
        return Response(
            data = products,
            status = 200
        )
    @action(detail=False, methods=['GET'])
    def get_products_filter(self, request):
        creator = request.GET.get('created_by')
        user_id = request.GET.get('user_id')
        condition = request.GET.get('condition')
        print("frima", condition)
        superuser = self.isSuperUser(user_id)

        products = []

        if creator == '':
            if superuser == True:
                print("super")
                products = Product.objects.filter(Q(ebay_url__icontains=condition) | Q(purchase_url__icontains=condition) | Q(product_name__icontains=condition),deleted = False).select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'notes', 'created_by_id')
                
        else:
            products = Product.objects.filter(Q(ebay_url__icontains=condition) | Q(purchase_url__icontains=condition) | Q(product_name__icontains=condition), created_by=creator, deleted=False).select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'notes', 'created_by_id')

        return Response(
            data = products,
            status = 200
        )
    
    @action(detail=False, methods=['POST'])
    def add_item(self, request):
        item = request.data['product']
        mode = request.data['mode']

        now = datetime.datetime.now()
        created_at = now.strftime("%Y-%m-%d")

        if mode == 1:
            try:
                # products = Product.objects.filter(Q(purchase_url = item['purchase_url']) | Q(ebay_url = request.data['ebay_url'])).values()
                products = Product.objects.filter(purchase_url = item['purchase_url'], deleted = False).values()


            except Exception as err:

                return Response(
                    data = 'データベース接続失敗！',
                    status = 401
                )

            if len(products) > 0:
                return Response(
                    data = 'すでに存在しています！',
                    status = 401
                )
            
            try:

                product = Product(
                    created_at = created_at,
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    created_by = request.user,
                    notes = item['notes']
                )
                
                product.save()

                return Response(
                    {'Success!'},
                    status = 200
                )
                
            except:
                return Response(
                    data = '登録操作が失敗しました',
                    status = 401
                )
        else:
            try:
                pid = item['id']

                product = Product.objects.filter(id = pid)
                product.update(
                    # created_at = created_at,
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    notes = item['notes']
                )

                return Response(
                    {'Success!'},
                    status = 200
                )
                
            except:
                return Response(
                    data = '編集操作が失敗しました！',
                    status = 401
                )
        
    @action(detail=False, methods=['GET'])  
    def get_results(self, request):

        worker = request.GET.get('worker')
        month = request.GET.get('month')

        if worker != "": 
            results = Product.objects.filter(created_by = worker, created_at__startswith = month).order_by('-id').values('id', 'created_at', 'created_by')
            return Response(data = results, status = 200)
        else:
            return Response(data = "", status = 200)


    @action(detail=False, methods=['POST'])
    def upload_product_file(self, request):
        file = request.FILES['csvFile']

        date = str(datetime.datetime.now()).split(" ")[0]

        wb = openpyxl.load_workbook(file)
        wproductsInfo = wb['Book1']

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        print("test rate", rate)
        res['rate'] = float(rate)

        for i in range(1, wproductsInfo.max_row):
            purchase_url = wproductsInfo.cell(row=i+1, column=4).value

            if len(Product.objects.filter(purchase_url = purchase_url, deleted = False)) > 0:
                continue

            purchase_price = str(wproductsInfo.cell(row=i+1,column=8).value).replace("¥", "").replace(",", "")
            prima = str(wproductsInfo.cell(row=i+1,column=9).value).replace("¥", "").replace(",", "")
            sell_price_en = str(wproductsInfo.cell(row=i+1,column=11).value).replace("$", "").replace(",", "")
            shipping = str(wproductsInfo.cell(row=i+1,column=12).value).replace("$", "").replace(",", "")

            if purchase_url == None or purchase_url == "":
                break

            if purchase_price.isnumeric() == False or purchase_price == None:
                purchase_price = 0

            if prima.isnumeric() == False or prima == None:
                prima = 0

            if shipping.isnumeric() == False or shipping == None:
                shipping = 0

            profit = '%.3f'%(profit_formula(float(sell_price_en), int(purchase_price), float(prima), float(shipping), res))
            profit_rate = 0

            if sell_price_en != 0:
                profit_rate = float(profit) / (float(sell_price_en) * float(rate)) * 100.0

            product = Product(
                created_at = date,
                product_name = wproductsInfo.cell(row=i+1, column=2).value,
                ec_site = wproductsInfo.cell(row=i+1, column=3).value,
                purchase_url = purchase_url,
                ebay_url = wproductsInfo.cell(row=i+1, column=5).value,
                purchase_price = purchase_price,
                sell_price_en = sell_price_en,
                profit = profit,
                profit_rate = profit_rate,
                prima = prima,
                shipping = shipping,
                quantity = 0,
                created_by = request.user,
                notes = wproductsInfo.cell(row=i+1, column=13).value
            )

            product.save()

        return Response(
            data="success",
            status=200
        )

    @action(detail=False, methods=['POST'])
    def upload_order_file(self, request):
        file = request.FILES['csvFile']

        date = str(datetime.datetime.now()).split(" ")[0]

        wb = openpyxl.load_workbook(file)
        wproductsInfo = wb['Book1']

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        print("test rate", rate)
        res['rate'] = str(rate)

        for i in range(1, wproductsInfo.max_row):
            purchase_url = wproductsInfo.cell(row=i+1, column=4).value

            if len(OrderList.objects.filter(purchase_url = purchase_url)) > 0:
                continue

            
            purchase_price = str(wproductsInfo.cell(row=i+1,column=8).value).replace("¥", "").replace(",", "")
            prima = str(wproductsInfo.cell(row=i+1,column=9).value).replace("¥", "").replace(",", "")
            sell_price_en = str(wproductsInfo.cell(row=i+1,column=11).value).replace("$", "").replace(",", "")
            shipping = str(wproductsInfo.cell(row=i+1,column=12).value).replace("$", "").replace(",", "")

            if purchase_url == None or purchase_url == "":
                break

            if purchase_price.isnumeric() == False or purchase_price == None:
                purchase_price = 0

            if prima.isnumeric() == False or prima == None:
                prima = 0

            if shipping.isnumeric() == False or shipping == None:
                shipping = 0

            profit = '%.3f'%(profit_order_formula(float(sell_price_en), int(purchase_price), float(prima), float(shipping), res))
            profit_rate = 0

            if sell_price_en != 0:
                profit_rate = float(profit) / (float(sell_price_en) * float(rate)) * 100.0


            product = OrderList(
                created_at = date,
                product_name = wproductsInfo.cell(row=i+1, column=2).value,
                ec_site = wproductsInfo.cell(row=i+1, column=3).value,
                purchase_url = purchase_url,
                ebay_url = wproductsInfo.cell(row=i+1, column=5).value,
                purchase_price = purchase_price,
                sell_price_en = sell_price_en,
                profit = profit,
                profit_rate = profit_rate,
                prima = prima,
                shipping = shipping,
                quantity = 0,
                created_by = request.user,
                notes = wproductsInfo.cell(row=i+1, column=13).value
            )

            product.save()

        return Response(
            data="success",
            status=200
        )
    
        
    @action(detail=False, methods=['GET'])    
    def get_orders(self, request):
        orders = OrderList.objects.select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'order_num','ordered_at', 'notes', 'created_by_id')
        print("test start",orders[0])
        return Response(
            data = orders,
            status = 200
        )
    @action(detail=False, methods=['GET'])    
    def get_filter_orders(self, request):
        # formatted_Date= request.data['formatted_Date']
        try :
            formatted_Date = request.GET.get('formatted_Date')
        
            # orders = OrderList.objects.select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'order_num', 'notes', 'created_by_id')
            orders = OrderList.objects.filter(
                created_at__startswith=formatted_Date
            ).order_by('-id').values('id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'order_num', 'notes', 'created_by_id')
            # print("test order", orders[0])
            if orders == None :
                return Response(
                    data = 'マッチするデータはありません。！',
                        status = 401
                )
            return Response(
                data = orders,
                status = 200
            )
        except :
            return Response(
                data = 'マッチするデータはありません。！',
                    status = 401
            )
    @action(detail=False, methods=['POST'])
    def add_order_item(self, request):
        item = request.data['order']
        mode = request.data['mode']
        date = datetime.datetime.now()
        if mode == 1:
            try:
                order = OrderList(
                    created_at = date,
                    product_name = item['product_name'],
                    ec_site = item['ec_site'],
                    purchase_url = item['purchase_url'],
                    ebay_url = item['ebay_url'],
                    purchase_price = item['purchase_price'],
                    sell_price_en = item['sell_price_en'],
                    profit = item['profit'],
                    profit_rate = item['profit_rate'],
                    prima = item['prima'],
                    shipping = item['shipping'],
                    quantity = item['quantity'],
                    order_num = item['order_num'],
                    ordered_at = request.user,
                    created_by = request.user,
                    notes = item['notes']
                )
                
                order.save()

                return Response(
                    {'Success!'},
                    status=200
                )
            
            except:
                return Response(
                    data = 'オーダー商品登録作業が失敗しました！',
                    status = 401
                )
        else:
            try:
                order_id = item['id']


                print(order_id)

                order = OrderList.objects.get(pk=order_id)

                print(order.product_name)
                date_string = str(item['created_at'])[:10]
                # order.update(
                #     created_at = item['created_at'],
                #     product_name = item['product_name'],
                #     ec_site = item['ec_site'],
                #     purchase_url = item['purchase_url'],
                #     ebay_url = item['ebay_url'],
                #     purchase_price = item['purchase_price'],
                #     sell_price_en = item['sell_price_en'],
                #     profit = float(item['profit']),
                #     profit_rate =float(item['profit_rate']),
                #     prima = item['prima'],
                #     shipping = item['shipping'],
                #     # quantity = item['quantity'],
                #     order_num = item['order_num'],
                #     notes = item['notes']
                # )
                print('time', item['created_at'])
                order.created_at =date_string + str(datetime.datetime.now().time())
                order.product_name = item['product_name']
                order.ec_site = item['ec_site']
                order.purchase_url = item['purchase_url']
                order.ebay_url = item['ebay_url']
                order.purchase_price = item['purchase_price']
                order.sell_price_en = item['sell_price_en']
                order.profit = item['profit']
                order.profit_rate = item['profit_rate']
                order.prima = item['prima']
                order.shipping = item['shipping']
                order.quantity = 999
                order.order_num = item['order_num']
                order.notes = item['notes']
                order.ordered_at = item['ordered_at']
                order.save()
                return Response(
                    {'Success!'},
                    status=200
                )
            
            except:
                return Response(
                    data = 'オーダー商品登録作業が失敗しました！',
                    status = 401
                )

    @action(detail=False, methods=['GET'])  
    def get_deleted_items(self, request):

        deletes_list = DeletedList.objects.select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima', 'shipping', 'notes', 'deleted_at', 'created_by_id')

        return Response(
            data = deletes_list,
            status = 200
        )
    @action(detail=False, methods=['GET'])
    def get_custom_deleted_products(self, request):
        

        products = []

        
        products = DeletedList.objects.select_related('created_by').order_by('-id').values('created_by__username', 'id', 'created_at', 'product_name', 'ec_site', 'purchase_url', 'ebay_url', 'purchase_price', 'sell_price_en', 'profit', 'profit_rate', 'prima','deleted_at', 'shipping', 'notes', 'created_by_id')
        
        return Response(
            data = products,
            status = 200
        )
    
    @action(detail=False, methods=['POST'])  
    def delete_product(self, request):
        id = request.data['id']
        
        item = Product.objects.filter(id = id)[0]

        try:
            date = datetime.datetime.now()
            del_item = ()

            del_item = DeletedList(
                    created_at = item.created_at,
                    updated_at = "",
                    product_name = item.product_name,
                    ec_site = item.ec_site,
                    purchase_url = item.purchase_url,
                    ebay_url = item.ebay_url,
                    purchase_price = item.purchase_price,
                    sell_price_en = item.sell_price_en,
                    profit = item.profit,
                    profit_rate = item.profit_rate,
                    prima = item.prima,
                    shipping = item.shipping,
                    quantity = item.quantity,
                    created_by = item.created_by,
                    notes = item.notes,
                    deleted_at = date
                )
            
            del_item.save()
            Product.objects.filter(id = id).update(deleted = True)

            return Response(data='success!', status=200)
        except:
            return Response(
                data = '削除操作が失敗しました！',
                status = 401
            )
    @action(detail=False, methods=['POST'])  
    def migrate_product(self, request):
        id = request.data['id']
        user= request.data['user']
        if user is None:
                user = User.objects.get_or_create(username='root')[0]  # Get or create user with username 'root'
        else:
            user = User.objects.get(username=user)
        item = Product.objects.filter(id = id)[0]
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        print("test rate", rate)
        res['rate'] = str(rate)
        new_profit = '%.3f'%(profit_order_formula(float(item.sell_price_en), int(item.purchase_price), float(item.prima), float(item.shipping), res))
        new_profit_rate = 0

        if float(item.sell_price_en) != 0:
            new_profit_rate = float(new_profit) / (float(item.sell_price_en) * float(rate)) * 100.0

        try:
            date = datetime.datetime.now()
            

            order = OrderList(
                    created_at = date,
                    updated_at = "",
                    product_name = item.product_name,
                    ec_site = item.ec_site,
                    purchase_url = item.purchase_url,
                    ebay_url = item.ebay_url,
                    purchase_price = item.purchase_price,
                    sell_price_en = item.sell_price_en,
                    profit = new_profit,
                    profit_rate = new_profit_rate,
                    prima = item.prima,
                    shipping = item.shipping,
                    quantity = item.quantity,
                    order_num = '',
                    created_by=user,
                    ordered_at=user.username,
                    notes = item.notes,
                )
            
            order.save()
            Product.objects.filter(id = id).update(deleted = True)
            return Response(
                {'Success!'},
                status=200
            )
            
        except:
            return Response(
                data = 'オーダー商品登録作業が失敗しました！',
                status = 401
            )
    @action(detail=False, methods=['POST'])  
    def migrate_del_item(self, request):
        try:
           
            id = request.data['id']
                       
            # Handling the case where user is None
            user = request.data['user']
            if user is None:
                user = User.objects.get_or_create(username='root')[0]  # Get or create user with username 'root'
            else:
                user = User.objects.get(username=user)
                
           

            item = DeletedList.objects.filter(id=id).first()
            if not item:
                return Response(data='アイテムが見つかりません', status=404)

            date = datetime.datetime.now()

            with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'), mode='r', encoding='utf-8') as f:
                settings_attrs = f.read()

            res = json.loads(settings_attrs)
            rate = getCurrentRate('JPY')
            print("test rate", rate)
            res['rate'] = str(rate)
            new_profit = '%.3f' % (profit_order_formula(float(item.sell_price_en), int(item.purchase_price), float(item.prima), float(item.shipping), res))
            new_profit_rate = 0

            if float(item.sell_price_en) != 0:
                new_profit_rate = float(new_profit) / (float(item.sell_price_en) * float(rate)) * 100.0

            order = OrderList(
                created_at=date,
                updated_at="",
                product_name=item.product_name,
                ec_site=item.ec_site,
                purchase_url=item.purchase_url,
                ebay_url=item.ebay_url,
                purchase_price=item.purchase_price,
                sell_price_en=item.sell_price_en,
                profit=new_profit,
                profit_rate=new_profit_rate,
                prima=item.prima,
                shipping=item.shipping,
                quantity=0,
                order_num='',
                created_by=user,
                ordered_at=user.username,  # Assuming ordered_at should be a string representing the user's username
                notes=item.notes,
            )

            order.save()
            
            # DeletedList.objects.get(id=item.id).delete()
            # Product.objects.filter(id=id).update(deleted=True)
            return Response(
                data='オーダー商品登録作業が成功しました！',
                status=200
            )

        except Exception as e:
            print(f"Exception occurred: {e}")
            return Response(
                data=f'オーダー商品登録作業が失敗しました: {str(e)}',
                status=401
            )
    @action(detail=False, methods=['POST'])  
    def recover_del_item(self, request):
        id = request.data['id']
        item = DeletedList.objects.filter(id = id)[0]
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'productmanage.settings')
  
        env = environ.Env()
        BASE_DIR = Path(__file__).resolve().parent
        env.read_env(str(BASE_DIR / ".env"))
        conn = psycopg2.connect(
                database = env('DB_NAME'),
                host = env('DB_HOST'),
                user = env('DB_USER'),
                password = env('DB_PASSWORD'),
                port = env('DB_PORT')
            )
            
        sql = "SELECT email, app_id, cert_id, dev_id, ebay_token FROM users_user WHERE is_superuser = TRUE"

        cur = conn.cursor()
        cur.execute(sql)
        row = cur.fetchone()

        if row == None:
            return Response(
                data = '商品再登録が失敗しました！',
                status = 401
            )

        
        ebay_setting = {
            'app_id' : row[1],
            'cert_id' : row[2],
            'dev_id' : row[3],
            'ebay_token' : row[4]
        }
        
        
        
        # print(request.user,'del mig', item.product_name, item.ec_site, item.purchase_url, item.ebay_url, item.purchase_price, item.sell_price_en, 
        # item.profit, item.profit_rate, item.prima, item.shipping, item.notes)
        
        try:
            
            
            if item.ebay_url == '':
                return Response(
                    data = '商品再登録が失敗しました！',
                    status = 401
                )
            Product.objects.filter(ebay_url = item.ebay_url).update(deleted = False)
            DeletedList.objects.get(id = id).delete()
            if ebay_setting['app_id'] == "" or ebay_setting['cert_id'] == "" or ebay_setting['dev_id'] == "" or ebay_setting['ebay_token'] == "":
                return Response(
                    data = '商品再登録が失敗しました！',
                    status = 401
                )
            
            item_number = item.ebay_url.split("/")[-1]

            try:
                api = Connection(appid = ebay_setting['app_id'], devid = ebay_setting['dev_id'], certid = ebay_setting['cert_id'], token = ebay_setting['ebay_token'], config_file=None)
                item_ebay = {
                    'Item': {
                        'ItemID': item_number.strip(),
                        'Quantity': 1
                    }
                }

                try:
                    api.execute('ReviseItem', item_ebay)
                    
                    
                    return Response(
                        data = '商品再登録が成功しました！',
                        status=200
                    )
                
                except:
                    print("api revise error")
                    return Response(
                        data = '商品再登録が失敗しました！',
                        status = 401
                    )
            
            except:
                print("api error")
                return Response(
                    data = '商品再登録が失敗しました！',
                    status = 401
                )
            
            
        except:
            return Response(
                data = '商品再登録が失敗しました！',
                status = 401
            )
        
    @action(detail=False, methods=['POST'])  
    def set_ebay_price(self, request):
        price = 0.00
        profitSellRate = request.data['profitSellRate']
        # item_url = request.data['ebay_url']
        # item = DeletedList.objects.filter(id = id)[0]
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        res['rate'] = str(rate)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(res, indent=4))
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'productmanage.settings')
  
        env = environ.Env()
        BASE_DIR = Path(__file__).resolve().parent
        env.read_env(str(BASE_DIR / ".env"))
        conn = psycopg2.connect(
                database = env('DB_NAME'),
                host = env('DB_HOST'),
                user = env('DB_USER'),
                password = env('DB_PASSWORD'),
                port = env('DB_PORT')
            )
            
        sql = "SELECT email, app_id, cert_id, dev_id, ebay_token FROM users_user WHERE is_superuser = TRUE"

        cur = conn.cursor()
        cur.execute(sql)
        row = cur.fetchone()

        if row == None:
            return Response(
                data = '商品再登録が失敗しました！',
                status = 401
            )

        
        ebay_setting = {
            'app_id' : row[1],
            'cert_id' : row[2],
            'dev_id' : row[3],
            'ebay_token' : row[4]
        }
        
        try:
            # update profits
            products = Product.objects.all().order_by('-id')
            ct=0
            for product in products:
                # sell_price = float(product.sell_price_en)
                if product.deleted == False:
                    purchase_price = product.purchase_price
                    prima = product.prima
                    shipping = product.shipping

                    price = float(sell_set_formula(float(profitSellRate), int(purchase_price), float(prima), float(shipping), res))
                    # print("test price1", price, rate, profitSellRate,  item_url, product.ebay_url)
                    profit = float(price)*float(rate)*float(profitSellRate)/100
                    # print("test price2", price, profit, item_url, product.ebay_url)
                    # profit_rate = (profit / (sell_price * rate)) * 100
                    
                    if ebay_setting['app_id'] == "" or ebay_setting['cert_id'] == "" or ebay_setting['dev_id'] == "" or ebay_setting['ebay_token'] == "":
                        return Response(
                            data = '商品再登録が失敗しました！',
                            status = 401
                        )
                    
                    item_number = product.ebay_url.split("/")[-1]
                    # print("test item", item_number," ", price)
                    try:
                        api = Connection(appid = ebay_setting['app_id'], devid = ebay_setting['dev_id'], certid = ebay_setting['cert_id'], token = ebay_setting['ebay_token'], config_file=None)
                        item_ebay = {
                            'Item': {
                                'ItemID': item_number.strip(),
                                'StartPrice': price
                            }
                        }
                        # print("test item", item_number," ", price, profit, profitSellRate)
                        
                        try:
                            api.execute('ReviseItem', item_ebay)
                            
                            Product.objects.filter(id = product.id).update(profit = profit, profit_rate = profitSellRate, sell_price_en = price)
                            
                            ct+=1
                            if ct>100 :
                                ct=0
                                print("test ct", ct)
                                for x in range(500000000):
                                    if x == 49999999:
                                        
                                        print("test x")
                        except:
                            print(product.purchase_url, "api revise error 1",product.ebay_url)
                            print("test ct error", ct)
                            
                                # return Response(
                                #     data = '商品再登録が失敗しました！',
                                #     status = 401
                                # )
                        
                        
                    except:
                        print("api error 2")
                        return Response(
                            data = '商品再登録が失敗しました！',
                            status = 401
                        )
                    

            return Response(
                data = '商品再登録が成功しました！',
                status=200
            )   
        except:
            print("api error 3")
            return Response(
                data = "商品再登録が失敗しました！",
                status = 401
            )
        
        
        
    @action(detail=False, methods=['POST'])
    def delete_order_item(self, request):
        pid = request.data['id']

        try:
            OrderList.objects.get(id = pid).delete()

            return Response(data='success!', status=200)
        except:
            return Response(
                data = '削除操作が失敗しました！',
                status = 401
            )
    
    @action(detail=False, methods=['GET'])
    def shipping_fee(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/shipping_fee.txt'),  mode='r', encoding='utf-8') as f:
            fee = f.read()
        
        return Response(
            data=json.loads(fee),
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def settings_attr(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        print("test rate", rate)
        
        res['rate'] = str(rate)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(res, indent=4))
        
        return Response(
            data = res,
            status = 200
        )
    
    @action(detail=False, methods=['POST'])
    def update_settings_attr(self, request):
        settings_attr = request.data['settings_attr']

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(settings_attr, ensure_ascii=False, indent=4))

        return Response(
            data=settings_attr,
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def get_ecsites(self, request):
        print("test ec")
        with open(file=str(settings.BASE_DIR / 'utils/scrape_site.txt'),  mode='r', encoding='utf-8') as f:
            ecsites = f.read()
        
        return Response(
            data=json.loads(ecsites),
            status=200
        )
    
    @action(detail=False, methods=['POST'])
    def update_ecsites(self, request):
        ecsites = request.data['ecsites']
        with open(file=str(settings.BASE_DIR / 'utils/scrape_site.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(ecsites, indent=4))

        return Response(
            data='Success',
            status=200
        )

    
    @action(detail=False, methods=['POST'])
    def download_product(self, request):
        with open(file=str(settings.BASE_DIR / 'media/products.csv'),  mode='w', encoding='utf-8', newline='') as f:
            fieldnames = ['商品名', 'EC site', '仕入れ URL', 'eBay URL', '利益額', '利益率', '仕入価格（円）', 'フリマ送料', '仕入合計（円）', '販売価格', '輸出送料', '出品者', '備考']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for product in Product.objects.all():
                writer.writerow({
                    '商品名': product.product_name,
                    'EC site': product.ec_site, 
                    '仕入れ URL': product.purchase_url, 
                    'eBay URL': product.ebay_url, 
                    '利益額': product.profit, 
                    '利益率': product.profit_rate, 
                    '仕入価格（円）': product.purchase_price, 
                    'フリマ送料': product.prima, 
                    '仕入合計（円）': product.purchase_price + product.prima, 
                    '販売価格': product.sell_price_en, 
                    '輸出送料': product.shipping, 
                    '出品者': product.created_by, 
                    '備考': product.notes
                })
        return Response(
            data='Success',
            status=200
        )
    
    @action(detail=False, methods=['GET'])
    def update_info(self, request):
        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='r', encoding='utf-8') as f:
            settings_attrs = f.read()

        res = json.loads(settings_attrs)
        rate = getCurrentRate('JPY')
        print("test rate", rate)
        res['rate'] = str(rate)

        with open(file=str(settings.BASE_DIR / 'utils/settings_attrs.txt'),  mode='w', encoding='utf-8') as f:
            f.write(json.dumps(res, indent=4))

        try:
            # update profits
            products = Product.objects.all().order_by('-id')

            for product in products:
                sell_price = float(product.sell_price_en)
                purchase_price = product.purchase_price
                prima = product.prima
                shipping = product.shipping

                profit = profit_formula(float(sell_price), int(purchase_price), float(prima), float(shipping), res)
                profit_rate = (profit / (sell_price * rate)) * 100

                Product.objects.filter(id = product.id).update(profit = profit, profit_rate = profit_rate)

            return Response(
                data = res,
                status = 200
            )
        except:
            return Response(
                data = "",
                status = 401
            )
            

    @action(detail=False, methods=['POST'])
    def get_item_specific(self, request):
        item_number = request.data['item_number']

        try:
            # Set up the API connection
            api = Connection(appid=request.user.app_id, devid=request.user.dev_id, certid=request.user.cert_id, token=request.user.ebay_token, config_file=None)
            response = api.execute(
                'GetItem',
                {
                    'ItemID': item_number,
                    'IncludeItemSpecifics': 'True'
                }
            )
            item_specifics = response.reply.Item.ItemSpecifics.NameValueList
            required_specifics = []
            optional_specifics = []
            for item_specific in item_specifics:
                if(getattr(item_specific, 'Name') == 'Brand' or getattr(item_specific, 'Name') == 'Type'):
                    required_specifics.append(
                        {
                            'Name': item_specific.Name,
                            'Value': item_specific.Value,
                            'Condition': 'Required',
                        }
                    )
                else:
                    optional_specifics.append(
                        {
                            'Name': item_specific.Name,
                            'Value': item_specific.Value,
                            'Condition': 'Optional',
                        }
                    )
            result = required_specifics + optional_specifics
            return Response(
                data = result,
                status = 200
            )
        except Exception as err:
            raise err
        
    def isSuperUser(self, user_id):
        user = User.objects.filter(id = user_id).values()
        return user[0]['is_superuser']

            
